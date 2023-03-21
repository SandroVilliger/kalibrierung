"""
Programm für die Kalibrierung von Kraftsensoren.
Es wird im Programm zusätzlich noch der Prüfbericht erstellt. Es werden dafür die Daten ins Excel exportiert.

Programm: Sandro Villiger
"""

#%% Import der Pakete

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.dataframe import dataframe_to_rows


#%% Veränderliche Variablen

# Pfad zu Datei mit den Rohdaten

file = r'C:\Users\Sandro Villiger\switchdrive\Arbeit IBI\Arbeit Uwe\037 Kallibrirung Killian\Programm\Testlauf_2_Kraftgeregelt.XLSX'

# Maximale Last in Newton (Absolutwert)

F_max = 2000

# Messstrom

S=2

# Abweichung Referenznormal

Urn= 0.022

# Abweichung Messverstärker

Umv = 0.05

# Bestimmung Zug- oder Druckversuch

'''
Zugverscuch = 1
Druckversuch = 0
'''

Versuch = 0

if Versuch == 0:
    F_max = F_max * -1
else:
    F_max = F_max

#%% Bestimmte Variablen

Stufen = np.array([0.125, 0.25, 0.375, 0.5, 0.675, 0.75, 0.875, 1])
Laststufen = F_max * Stufen

# Verdoplung der Abweichung
k=2

#%% Import der Daten, entfernen der unnötigen Daten und einlesen in Numpy

# Header Name

name = ['Zeit', 'Referenz kN', 'Prüfling kN']

# Import der Daten

data = pd.read_excel(file, header=None, names=name)

# Löschen der unnötigen Daten

delet = np.arange(0, 49, 1)
Werte = data.drop(delet, axis=0)


# Einlesen der Daten in np.array und umrechnung in Newton

Zeit = (np.round(np.array(Werte['Zeit'], dtype='float64'), 1))
Referenz =( np.array(Werte['Referenz kN'], dtype='float64'))*1000
Prüfling = (np.array(Werte['Prüfling kN'], dtype='float64'))*1000

i=Zeit.reshape(2568,1)
i=np.append(i,Referenz.reshape(2568,1),axis=1)
Werte=pd.DataFrame(np.append(i,Prüfling.reshape(2568,1),axis=1),columns=name)




#%% Plot der gesamten Daten

y_Beschriftung = np.linspace(0, F_max, num=5)
'''
plt.figure(figsize=(8, 5), dpi=300)
plt.plot(Zeit,Referenz,color='black')
plt.hlines(0,min(Zeit),max(Zeit),colors='grey')
plt.vlines(min(Zeit),0,F_max,colors='grey')
plt.title('Gesamt Daten', pad=12,fontsize=12)
plt.xlabel('Zeit [s]' ,fontsize=12,)
plt.ylabel('Kraft [kN]' ,fontsize=12,)
plt.yticks(y_Beschriftung)
plt.savefig('Gesamte Messung.png')
plt.show()'''

#%% Slicing der Daten nach den Laststufen


Mittelwerte_Refernz = [[], [], [], [], [], [], [],  []]

Mittelwerte_Prüfling = [[], [], [], [], [], [], [], []]

Werte_F = Werte[(Werte['Referenz kN'] >= 0- 10) & (Werte['Referenz kN'] <=0 + 10)]['Referenz kN'].mean()

for x in range(len(Laststufen)):
    Werte_F = Werte[(Werte['Referenz kN'] >= Laststufen[x] - 10) & (Werte['Referenz kN'] <= Laststufen[x] + 10)]
    Werte_F_index = Werte[(Werte['Referenz kN'] >= Laststufen[x] - 10) & (Werte['Referenz kN'] <= Laststufen[x] + 10)].index.values
    res = np.where(Werte_F_index[:-1] + 1 != Werte_F_index[1:])[0]
    res = res[res > 10]

    X_max_1 = Werte_F[10:res[0] -9]
    X_max_2 = Werte_F[res[0] + 10:res[1] -9]
    X_max_3 = Werte_F[res[1] + 10:res[2] -9]
    X_max_4 = Werte_F[res[2] + 10:-10]


    Mittelwerte_Refernz[x].append(X_max_1.mean()[1])
    Mittelwerte_Refernz[x].append(X_max_2.mean()[1])
    Mittelwerte_Refernz[x].append(X_max_3.mean()[1])
    Mittelwerte_Refernz[x].append(X_max_4.mean()[1])

    Mittelwerte_Prüfling[x].append(X_max_1.mean()[2])
    Mittelwerte_Prüfling[x].append(X_max_2.mean()[2])
    Mittelwerte_Prüfling[x].append(X_max_3.mean()[2])
    Mittelwerte_Prüfling[x].append(X_max_4.mean()[2])

    '''plt.figure(figsize=(8, 5), dpi=300)
    plt.plot(Zeit, Referenz, color='black')
    plt.plot(X_max_1['Zeit'], X_max_1['Referenz kN'], color='blue')
    plt.plot(X_max_2['Zeit'], X_max_2['Referenz kN'], color='red')
    plt.plot(X_max_3['Zeit'], X_max_3['Referenz kN'], color='blue')
    plt.plot(X_max_4['Zeit'], X_max_4['Referenz kN'], color='red')
    plt.vlines(min(Zeit), 0, F_max, colors='grey')
    plt.title('Versuchsablauf Referenznormal', pad=12, fontsize=12)
    plt.xlabel('Zeit [s]', fontsize=12, )
    plt.ylabel('Kraft [kN]', fontsize=12, )
    plt.yticks(y_Beschriftung)
    plt.show()'''

Mittelwerte_Refernz = np.array(Mittelwerte_Refernz)
Mittelwerte_Prüfling = np.array(Mittelwerte_Prüfling)

#%% Auswertung

Refernz_mean = np.mean(Mittelwerte_Refernz, axis=1)
Prüfling_mean = np.mean(Mittelwerte_Prüfling, axis=1)

A = np.array(((Prüfling_mean-Refernz_mean ) / Laststufen) * 100)

# Abweichung

U=np.sqrt((Urn)**2+(Umv)**2)

# Abweichung in Newton

U_Newton= U/100*Laststufen

# Abweichung

Abweichung=Prüfling_mean-Refernz_mean

# Wiederholbarkeit

Wiederholbarkeit=np.std(Mittelwerte_Prüfling,axis=1,ddof=1)

# Hysterese

Hysterese1= abs(Mittelwerte_Prüfling[:,0]-Mittelwerte_Prüfling[:,2])
Hysterese2= abs(Mittelwerte_Prüfling[:,1]-Mittelwerte_Prüfling[:,3])
Hysterese=[]
for x in range(0,8,1):
    t=np.maximum(Hysterese1[x],Hysterese2[x])
    Hysterese.append(t)

Hysterese=np.array(Hysterese)

Genauigkeit=abs(Abweichung)+2*abs(U_Newton)

Messunsicherheit=np.concatenate((Refernz_mean.reshape(-1,1),Prüfling_mean.reshape(-1,1),Abweichung.reshape(-1,1),Wiederholbarkeit.reshape(-1,1),Hysterese.reshape(-1,1),U_Newton.reshape(-1,1),Genauigkeit.reshape(-1,1)),axis=1)

#%% Plot des Abweichungsdiagrams

text = ['F1', 'F2','F3', 'F4', 'F5', 'F6', 'F7','F8']

plt.figure(figsize=(8, 5), dpi=300)
plt.plot(abs(Laststufen),A,color='green')
plt.plot(abs(Laststufen),A-k*(U),color='blue',linestyle='--')
plt.plot(abs(Laststufen),A+k*(U),color='blue',linestyle='--')
plt.hlines(0.2,abs(Laststufen)[0],abs(Laststufen)[-1],colors='red')
plt.hlines(-0.2,abs(Laststufen)[0],abs(Laststufen)[-1],colors='red')
plt.xticks(abs(Laststufen), text)
plt.grid()
plt.savefig('Abweichung.png')
'''plt.show()'''

#%% Korrekturfaktor

Korrektur_max= (S/ ((F_max)-(A[-1]/100)*F_max))*F_max

Korrektur_unten=((A[0]/100)*F_max)+(F_max*Stufen[0])

Korrektur_min=(Stufen[0]*S*F_max/Korrektur_unten)-2

print(0,'N','=', Korrektur_min,'mV/V')
print(-F_max,'N','=', Korrektur_max,'mV/V')


#%% Excel Export



wb=Workbook()

# Defieneren der Schrift

ft = Font(name='Verdana',size=10) # Standard
ft2 = Font(name='Verdana',size=8) # Bildbeschriftung
ft3 = Font(name='Verdana',size=10,bold=True) # Hervorgehoben

# Erstellen des Worksheets
ws=wb.active

# Grösse für der Zelle (Richtige grösse für das Drucken)
ws.column_dimensions['G'].width=10
ws.column_dimensions['B'].width=12
ws.column_dimensions['C'].width=13
ws.column_dimensions['D'].width=17
ws.column_dimensions['E'].width=13
ws.column_dimensions['F'].width=16
ws.column_dimensions['G'].width=12

# Fusszeile
ws.oddFooter.left.text = "Seite &[Page] von &N"
ws.oddFooter.left.size = 11
ws.oddFooter.left.font = "Verdana"

# Einführen des Logos
img = Image('HSLU_Logo.png')
img.height = 30
img.width= 200
img.anchor = 'F1'
ws.add_image(img)
img = Image('HSLU_Logo.png')
img.height = 30
img.width= 200
img.anchor = 'F50'
ws.add_image(img)


# Titel des Dokuments
ws['A1']='Kalibrierschein Kraftsensoren'
ws['A1'].font=Font(bold=True, underline='single',name='Verdana',size=16)

# Text für Prüfbestimmungen

ws['A4']='Kalibriergegenstand:'
ws['A4'].font=ft
ws['A5']='Typ:'
ws['A5'].font=ft
ws['A6']='Log.Nr.(STS 0209):'
ws['A6'].font=ft
ws['A7']='Serien NR:'
ws['A7'].font=ft
ws['A8']='Anzeigebereich:'
ws['A8'].font=ft
ws['A9']='Max. zul. Abweichung:'
ws['A9'].font=ft
ws['A11']='Referenznormal:'
ws['A11'].font=ft
ws['A12']='Typ:'
ws['A12'].font=ft
ws['A13']='Log.Nr.:'
ws['A13'].font=ft
ws['A14']='Serien NR:'
ws['A14'].font=ft
ws['A16']='Datum der Kalibrierung:'
ws['A16'].font=ft
ws['A17']='Prüfer:'
ws['A17'].font=ft

ws['E5']='Temperatur:'
ws['E6']='Rel. Luftfeuchte:'
ws['E7']='Umgebungsdruck:'
ws['E8']='Ort der Kalibrierung:'
ws['E5'].font=ft
ws['E6'].font=ft
ws['E7'].font=ft
ws['E8'].font=ft

# Abgrenzung
Liste=('A','B','C','D','E','F','G',)
for i in Liste:
    ws[f'{i}17'].border=Border( bottom=Side(style='thin'))

ws['A20']='Die Kalibrierung erfolgt nach der Verfahrensanweisung BAA08082 "Kalibration Kraft" nach dem '
ws['A21']='Verfahren der DAkkS EA-4/02 M: 2013 "Ermittlung der Messunsicherheiten bei Kalibrierungen" '
ws['A20'].font=ft
ws['A21'].font=ft

# Text Mittelwerte
ws['A23']='Folgend werden die Mittelwerte über die verschieden Laststufen für den Prüfkörper gezeigt'
ws['A23'].font=ft

ws['A25']='Mittelwerte des Prüfkörpers in Newton'
ws['A25'].font=ft3

# Export der Mittelwerte

Mittelwerte_Prüfling_Export=pd.DataFrame(Mittelwerte_Prüfling,columns=['M1','M2','M3','M4'],index=(np.round(Refernz_mean,decimals=2)))

for r in dataframe_to_rows(np.round(Mittelwerte_Prüfling_Export,decimals=2),index=True,header=True):
    ws.append(r)
ws.move_range("A26:F26", rows=1, cols=0)


ws['B27'].font=Font(bold=True)
ws['B27'].fill=PatternFill("solid", fgColor="e6e6e6")
ws['C27'].font=Font(bold=True)
ws['C27'].fill=PatternFill("solid", fgColor="e6e6e6")
ws['D27'].font=Font(bold=True)
ws['D27'].fill=PatternFill("solid", fgColor="e6e6e6")
ws['E27'].font=Font(bold=True)
ws['E27'].fill=PatternFill("solid", fgColor="e6e6e6")
for i in range(27,36,1):
    ws[f'A{i}'].font=Font(bold=True)
    ws[f'A{i}'].fill=PatternFill("solid", fgColor="e6e6e6")

Liste1=('A','B','C','D','E',)

for L in Liste1:
    for i in range(27,36,1):
        ws[f'{L}{i}'].border=Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

ws['A37']='Messunsicherheit'
ws['A37'].font=ft3


Messunsicherheit_export=pd.DataFrame(Messunsicherheit)

for r in dataframe_to_rows(np.round(Messunsicherheit_export,decimals=2),index=True,header=True):
    ws.append(r)

for row in ws['A39:A47']:
    for cell in row:
        cell.value = None

for row in ws['A38:H38']:
    for cell in row:
        cell.value = None


ws.move_range("B40:H47", rows=0, cols=-1)


ws['A39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['A39'].value='Referenz\n[N]'
ws['A39'].font=Font(bold=True)
ws['A39'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['B39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['B39'].value='Mittelwert\n[N]'
ws['B39'].font=Font(bold=True)
ws['B39'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['C39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['C39'].value='Abweichung\n[N]'
ws['C39'].font=Font(bold=True)
ws['C39'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['D39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['D39'].value='Wiederholbarkeit\n[N]'
ws['D39'].font=Font(bold=True)
ws['D39'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['E39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['E39'].value='Hysterese\n[N]'
ws['E39'].font=Font(bold=True)
ws['E39'].fill=PatternFill("solid", fgColor="e6e6e6")

ws['F39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['F39'].value='Messunsicherheit\n[N]'
ws['F39'].font=Font(bold=True)
ws['F39'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['G39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['G39'].value='Genauigkeit\nSensor\n[N]'
ws['G39'].font=Font(bold=True,size=11)
ws['G39'].fill=PatternFill("solid", fgColor="e6e6e6")

Liste1=('A','B','C','D','E','F','G')

for L in Liste1:
    for i in range(39,48,1):
        ws[f'{L}{i}'].border=Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

ws['A53']='Die Abweichung des Referenznormals beträgt:'
ws['A53'].font=ft

ws['E53']=Urn
ws['E53'].font=ft3

ws['F53']='%'
ws['F53'].font=ft3

ws['A55']='Die Abweichung des Messverstärkers beträgt:'
ws['A55'].font=ft

ws['E55']=Umv
ws['E55'].font=ft3

ws['F55']='%'
ws['F55'].font=ft3

ws['A57']='Im der Abbildung 1 wird die Abweichung über die Laststufen in % dargestellt'
ws['A57'].font=ft

img = Image('Abweichung.png')
img.height = 375
img.width= 600
img.anchor = 'A59'
ws.add_image(img)

ws['B79']='Abbildung 2'
ws['B79'].font=ft2

ws['A81']='Die Genauigkeit des Kraftsensors beträgt:'
ws['A81'].font=ft3

ws['E81']=max(np.round(abs(Genauigkeit),decimals=2))
ws['E81'].font=ft3

ws['F81']='N'
ws['F81'].font=ft3

# Speichern des Excels

wb.save('Kalibrierung.xlsx')

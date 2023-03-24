"""
Programm für die Kalibrierung von Wasser. Es handelt sich um die Version 1.1.
Es wird im Programm zusätzlich noch der Prüfbericht erstellt. Es werden dafür die Daten ins Excel exportiert.

Programm: Sandro Villiger
"""

# %% Import der Pakete

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# %% Veränderliche Variablen

# Pfad zu Datei mit den Rohdaten

file = r'C:\Users\Sandro Villiger\switchdrive\Arbeit IBI\Arbeit Uwe\037 Kallibrirung Killian\Kallibrirung der Wassermenge\2022-12-22-10_50_42.475_Kalibration_Wasser_1.1.csv'

U = 0.09

# %% Bestimmte Variablen

k = 2

# %% Import der Daten


data = pd.read_csv(file, sep=';')

Referenz = np.array(data['MW'])

Messung = np.array(data['Qw 11'])

Punkte = np.arange(0, len(Referenz), 1)

# %% Plot der gesamten Daten

'''plt.figure(figsize=(8, 5), dpi=300)
plt.plot(Punkte, Referenz, color='black', label='Referenz')
plt.plot(Punkte, Referenz, color='blue', label='Messung')
plt.grid()
plt.xticks(np.arange(0, 10001, 1000))
plt.legend()
plt.show()'''

# %% Slicing der Daten nach Zeitschriten

# Referenz Daten

Max_1_1_2_r = Referenz[5292:5800]
Max_1_R = Max_1_1_2_r

Max_0875_1_1_r = Referenz[4741:5126]
Max_0875_1_2_r = Referenz[6050:6307]
Max_0875_R = np.concatenate((Max_0875_1_1_r, Max_0875_1_2_r), axis=0)

Max_075_1_1_r = Referenz[4300:4670]
Max_075_1_2_r = Referenz[6405:6853]
Max_075_R = np.concatenate((Max_075_1_1_r, Max_075_1_2_r), axis=0)

Max_0625_1_1_r = Referenz[3673:4126]
Max_0625_1_2_r = Referenz[7000:7398]
Max_0625_R = np.concatenate((Max_0625_1_1_r, Max_0625_1_2_r), axis=0)

Max_05_1_1_r = Referenz[3200:3580]
Max_05_1_2_r = Referenz[7549:7962]
Max_05_R = np.concatenate((Max_05_1_1_r, Max_05_1_2_r), axis=0)

Max_0375_1_1_r = Referenz[2606:3034]
Max_0375_1_2_r = Referenz[8098:8490]
Max_0375_R = np.concatenate((Max_0375_1_1_r, Max_0375_1_2_r), axis=0)

Max_025_1_1_r = Referenz[2254:2488]
Max_025_1_2_r = Referenz[8561:9035]
Max_025_R = np.concatenate((Max_025_1_1_r, Max_025_1_2_r), axis=0)

Max_0125_1_1_r = Referenz[1750:2050]
Max_0125_1_2_r = Referenz[9141:9581]
Max_0125_R = np.concatenate((Max_0125_1_1_r, Max_0125_1_2_r), axis=0)

# Prüfdaten Daten

Max_1_1_2_m = Messung[5292:5800]
Max_1_M = Max_1_1_2_m

Max_0875_1_1_m = Messung[4741:5126]
Max_0875_1_2_m = Messung[6050:6307]
Max_0875_M = np.concatenate((Max_0875_1_1_m, Max_0875_1_2_m), axis=0)

Max_075_1_1_m = Messung[4300:4670]
Max_075_1_2_m = Messung[6405:6853]
Max_075_M = np.concatenate((Max_075_1_1_m, Max_075_1_2_m), axis=0)

Max_0625_1_1_m = Messung[3673:4126]
Max_0625_1_2_m = Messung[7000:7398]
Max_0625_M = np.concatenate((Max_0625_1_1_m, Max_0625_1_2_m), axis=0)

Max_05_1_1_m = Messung[3200:3580]
Max_05_1_2_m = Messung[7549:7962]
Max_05_M = np.concatenate((Max_05_1_1_m, Max_05_1_2_m), axis=0)

Max_0375_1_1_m = Messung[2606:3034]
Max_0375_1_2_m = Messung[8098:8490]
Max_0375_M = np.concatenate((Max_0375_1_1_m, Max_0375_1_2_m), axis=0)

Max_025_1_1_m = Messung[2254:2488]
Max_025_1_2_m = Messung[8561:9035]
Max_025_M = np.concatenate((Max_025_1_1_m, Max_025_1_2_m), axis=0)

Max_0125_1_1_m = Messung[1750:2050]
Max_0125_1_2_m = Messung[9141:9581]
Max_0125_M = np.concatenate((Max_0125_1_1_m, Max_0125_1_2_m), axis=0)

# Punkte für Plot

Punkte_1 = np.arange(5292, 5800)
Punkte_875 = np.concatenate((np.arange(4741, 5126), np.arange(6050, 6307)), axis=0)
Punkte_75 = np.concatenate((np.arange(4300, 4670), np.arange(6405, 6853)), axis=0)
Punkte_625 = np.concatenate((np.arange(3673, 4126), np.arange(7000, 7398)), axis=0)
Punkte_05 = np.concatenate((np.arange(3200, 3580), np.arange(7549, 7962)), axis=0)
Punkte_375 = np.concatenate((np.arange(2606, 3034), np.arange(8098, 8490)), axis=0)
Punkte_25 = np.concatenate((np.arange(2254, 2488), np.arange(8561, 9035)), axis=0)
Punkte_125 = np.concatenate((np.arange(1750, 2050), np.arange(9141, 9581)), axis=0)

# %% Plot der Daten mit den verwendeten Messdaten
'''
plt.figure(figsize=(8, 5), dpi=300)
plt.plot(Punkte, Referenz, color='black', label='Referenz')
plt.plot(Punkte, Referenz, color='blue', label='Messung')

plt.plot(np.arange(5292, 5800), Max_1_1_2_r, color='red', label='Auswertung')
plt.plot(np.arange(4741, 5126), Max_0875_1_1_r, color='red')
plt.plot(np.arange(6050, 6307), Max_0875_1_2_r, color='red')

plt.plot(np.arange(4300, 4670), Max_075_1_1_r, color='red')
plt.plot(np.arange(6405, 6853), Max_075_1_2_r, color='red')

plt.plot(np.arange(3673, 4126), Max_0625_1_1_r, color='red')
plt.plot(np.arange(7000, 7398), Max_0625_1_2_r, color='red')

plt.plot(np.arange(3200, 3580), Max_05_1_1_r, color='red')
plt.plot(np.arange(7549, 7962), Max_05_1_2_r, color='red')

plt.plot(np.arange(2606, 3034), Max_0375_1_1_r, color='red')
plt.plot(np.arange(8098, 8490), Max_0375_1_2_r, color='red')

plt.plot(np.arange(2254, 2488), Max_025_1_1_r, color='red')
plt.plot(np.arange(8561, 9035), Max_025_1_2_r, color='red')

plt.plot(np.arange(1750, 2050), Max_0125_1_1_r, color='red')
plt.plot(np.arange(9141, 9581), Max_0125_1_2_r, color='red')

plt.grid()
plt.xticks(np.arange(0, 10001, 1000))
plt.legend()
plt.show()'''

# %% Berechnung der Mittelwerte

# Referenz

x1_R = Max_0125_1_1_r.mean()
x2_R = Max_025_1_1_r.mean()
x3_R = Max_0375_1_1_r.mean()
x4_R = Max_05_1_1_r.mean()
x5_R = Max_0625_1_1_r.mean()
x6_R = Max_075_1_1_r.mean()
x7_R = Max_0875_1_1_r.mean()
x8_R = Max_1_1_2_r.mean()
x9_R = Max_0875_1_2_r.mean()
x10_R = Max_075_1_2_r.mean()
x11_R = Max_0625_1_2_r.mean()
x12_R = Max_05_1_2_r.mean()
x13_R = Max_0375_1_2_r.mean()
x14_R = Max_025_1_2_r.mean()
x15_R = Max_0125_1_2_r.mean()


# Prüfdaten

x1_M = Max_0125_1_1_m.mean()
x2_M = Max_025_1_1_m.mean()
x3_M = Max_0375_1_1_m.mean()
x4_M = Max_05_1_1_m.mean()
x5_M = Max_0625_1_1_m.mean()
x6_M = Max_075_1_1_m.mean()
x7_M = Max_0875_1_1_m.mean()
x8_M = Max_1_1_2_m.mean()
x9_M = Max_0875_1_2_m.mean()
x10_M = Max_075_1_2_m.mean()
x11_M = Max_0625_1_2_m.mean()
x12_M = Max_05_1_2_m.mean()
x13_M = Max_0375_1_2_m.mean()
x14_M = Max_025_1_2_m.mean()
x15_M = Max_0125_1_2_m.mean()

M1=np.array([x2_M,x3_M,x4_M,x5_M,x6_M,x7_M,x8_M])
M2=np.array([x14_M,x13_M,x12_M,x11_M,x10_M,x9_M,np.nan])

X_Stufen = np.array([x1_M, x2_M, x3_M, x4_M, x5_M, x6_M, x7_M, x8_M, x9_M, x10_M, x11_M, x12_M, x13_M, x14_M, x15_M])

# %% Berechnung der totalen Mittelwerte


X1R = (x1_R + x15_R) / 2
X2R = (x2_R + x14_R) / 2
X3R = (x3_R + x13_R) / 2
X4R = (x4_R + x12_R) / 2
X5R = (x5_R + x11_R) / 2
X6R = (x6_R + x10_R) / 2
X7R = (x7_R + x9_R) / 2
X8R = x8_R
XR = np.array([X1R, X2R, X3R, X4R, X5R, X6R, X7R, X8R])

X1M = (x1_M + x15_M) / 2
X2M = (x2_M + x14_M) / 2
X3M = (x3_M + x13_M) / 2
X4M = (x4_M + x12_M) / 2
X5M = (x5_M + x11_M) / 2
X6M = (x6_M + x10_M) / 2
X7M = (x7_M + x9_M) / 2
X8M = x8_M

XM = np.array([X1M, X2M, X3M, X4M, X5M, X6M, X7M, X8M])

# %% Auswertung

Laststufen = np.array([1, 5, 10, 15, 20, 25, 30, 35])
print(XM - XR)
A = ((XM - XR) / Laststufen) * 100

U_Wasser= U/100*Laststufen[1:]

Abweichung=XM[1:] - XR[1:]

Hysterese= abs(M1-M2)

Genauigkeit=abs(Abweichung)+2*abs(U_Wasser)

Messunsicherheit=np.concatenate((XR[1:].reshape(-1,1),XM[1:].reshape(-1,1),Abweichung.reshape(-1,1),Hysterese.reshape(-1,1),U_Wasser.reshape(-1,1),Genauigkeit.reshape(-1,1)),axis=1)


# %% Plot Auswertungsdiagramm

text=['-10%','-7.5%','-5%','-2.5%','0%','2.5%','5%','7.5%','10%']

plt.figure(figsize=(8, 5), dpi=300)
plt.plot(abs(Laststufen[1:]),A[1:],color='green')
plt.plot(abs(Laststufen[1:]),A[1:]-k*(U),color='blue',linestyle='--')
plt.plot(abs(Laststufen[1:]),A[1:]+k*(U),color='blue',linestyle='--')
plt.hlines(0,abs(Laststufen)[1],abs(Laststufen)[-1],colors='black')
plt.hlines(10,abs(Laststufen)[1],abs(Laststufen)[-1],colors='red')
plt.hlines(-10,abs(Laststufen)[1],abs(Laststufen)[-1],colors='red')
plt.yticks(np.arange(-10,11,2.5), text)
plt.grid()
plt.savefig('Abweichung_1.1.png')
'''plt.show()'''




#%% Excel Export

wb=Workbook()

# Defieneren der Schrift

ft = Font(name='Verdana',size=10) # Standard
ft2 = Font(name='Verdana',size=8) # Bildbeschriftung
ft3 = Font(name='Verdana',size=10,bold=True) # Hervorgehoben

# Erstellen des Worksheets
ws=wb.active

# Grösse für der Zelle (Richtige grösse für das Drucken)
ws.column_dimensions['G'].width=10
ws.column_dimensions['B'].width=12
ws.column_dimensions['C'].width=13
ws.column_dimensions['D'].width=13
ws.column_dimensions['E'].width=17
ws.column_dimensions['F'].width=17
ws.column_dimensions['G'].width=13

# Fusszeile
ws.oddFooter.left.text = "Seite &[Page] von &N"
ws.oddFooter.left.size = 11
ws.oddFooter.left.font = "Verdana"

# Einführen des Logos
img = Image('HSLU_Logo.png')
img.height = 30
img.width= 200
img.anchor = 'F1'
ws.add_image(img)
img = Image('HSLU_Logo.png')
img.height = 30
img.width= 200
img.anchor = 'F50'
ws.add_image(img)


# Titel des Dokuments
ws['A1']='Kalibrierschein Wasser 1.1'
ws['A1'].font=Font(bold=True, underline='single',name='Verdana',size=16)

# Text für Prüfbestimmungen

ws['A4']='Kalibriergegenstand:'
ws['A4'].font=ft
ws['A5']='Typ:'
ws['A5'].font=ft
ws['A6']='Log.Nr.(STS 0209):'
ws['A6'].font=ft
ws['A7']='Serien NR:'
ws['A7'].font=ft
ws['A8']='Anzeigebereich:'
ws['A8'].font=ft
ws['A9']='Max. zul. Abweichung:'
ws['A9'].font=ft
ws['A11']='Referenznormal:'
ws['A11'].font=ft
ws['A12']='Typ:'
ws['A12'].font=ft
ws['A13']='Log.Nr.:'
ws['A13'].font=ft
ws['A14']='Serien NR:'
ws['A14'].font=ft
ws['A16']='Datum der Kalibrierung:'
ws['A16'].font=ft
ws['A17']='Prüfer:'
ws['A17'].font=ft

ws['E5']='Temperatur:'
ws['E6']='Rel. Luftfeuchte:'
ws['E7']='Umgebungsdruck:'
ws['E8']='Ort der Kalibrierung:'
ws['E5'].font=ft
ws['E6'].font=ft
ws['E7'].font=ft
ws['E8'].font=ft


# Abgrenzung
Liste=('A','B','C','D','E','F','G',)
for i in Liste:
    ws[f'{i}17'].border=Border( bottom=Side(style='thin'))

ws['A20']='Die Kalibrierung erfolgt nach der Verfahrensanweisung BAA08083 "Wasser" nach dem '
ws['A21']='Verfahren der DAkkS EA-4/02 M: 2013 "Ermittlung der Messunsicherheiten bei Kalibrierungen" '
ws['A20'].font=ft
ws['A21'].font=ft

# Text Mittelwerte
ws['A23']='Folgend werden die Mittelwerte über die verschieden Stufen für den Prüfkörper gezeigt'
ws['A23'].font=ft

ws['A25']='Mittelwerte des Prüfkörpers in Liter pro Minute'
ws['A25'].font=ft3


Tabelle_1=np.concatenate((XR[1:].reshape(-1,1),M1.reshape(-1,1),M2.reshape(-1,1)),axis=1)
Tabelle_1_Export=pd.DataFrame(np.round(Tabelle_1,decimals=2))
for r in dataframe_to_rows(Tabelle_1_Export):
    ws.append(r)

for row in ws['A26:A35']:
    for cell in row:
        cell.value = None

for row in ws['A26:H26']:
    for cell in row:
        cell.value = None

ws.move_range("B28:D35", rows=0, cols=-1)

for i in range(28,35,1):
    ws[f'A{i}'].font=Font(bold=True)
    ws[f'A{i}'].fill=PatternFill("solid", fgColor="e6e6e6")

ws['A27'].alignment = Alignment(wrap_text=True,vertical="top")
ws['A27'].value='Referenz'
ws['A27'].font=Font(bold=True)
ws['A27'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['B27'].alignment = Alignment(wrap_text=True,vertical="top")
ws['B27'].value='M1'
ws['B27'].font=Font(bold=True)
ws['B27'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['C27'].alignment = Alignment(wrap_text=True,vertical="top")
ws['C27'].value='M2'
ws['C27'].font=Font(bold=True)
ws['C27'].fill=PatternFill("solid", fgColor="e6e6e6")

Liste1=('A','B','C')

for L in Liste1:
    for i in range(27,35,1):
        ws[f'{L}{i}'].border=Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


Messunsicherheit_export=pd.DataFrame(Messunsicherheit)

for r in dataframe_to_rows(np.round(Messunsicherheit_export,decimals=2),index=True,header=True):
    ws.append(r)

ws['A36']='Messunsicherheit'
ws['A36'].font=ft3


for row in ws['A38:A44']:
    for cell in row:
        cell.value = None

for row in ws['B36:G36']:
    for cell in row:
        cell.value = None

ws.move_range("B38:G44", rows=1, cols=-1)

ws['A38'].alignment = Alignment(wrap_text=True,vertical="top")
ws['A38'].value='Referenz\n[L/Min]'
ws['A38'].font=Font(bold=True)
ws['A38'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['B38'].alignment = Alignment(wrap_text=True,vertical="top")
ws['B38'].value='Mittelwert\n[L/Min]'
ws['B38'].font=Font(bold=True)
ws['B38'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['C38'].alignment = Alignment(wrap_text=True,vertical="top")
ws['C38'].value='Abweichung\n[L/Min]'
ws['C38'].font=Font(bold=True)
ws['C38'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['D38'].alignment = Alignment(wrap_text=True,vertical="top")
ws['D38'].value='Hysterese\n[L/Min]'
ws['D38'].font=Font(bold=True)
ws['D38'].fill=PatternFill("solid", fgColor="e6e6e6")

ws['E38'].alignment = Alignment(wrap_text=True,vertical="top")
ws['E38'].value='Messunsicherheit\n[L/Min]'
ws['E38'].font=Font(bold=True)
ws['E38'].fill=PatternFill("solid", fgColor="e6e6e6")

ws['F38'].alignment = Alignment(wrap_text=True,vertical="top")
ws['F38'].value='Genauigkeit\nSensor\n[L/Min]'
ws['F38'].font=Font(bold=True,size=11)
ws['F38'].fill=PatternFill("solid", fgColor="e6e6e6")

Liste1=('A','B','C','D','E','F')

for L in Liste1:
    for i in range(38,46,1):
        ws[f'{L}{i}'].border=Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


ws['A53']='Die Abweichung des Referenznormals beträgt:'
ws['A53'].font=ft

ws['E53']=U
ws['E53'].font=ft3

ws['F53']='%'
ws['F53'].font=ft3

ws['A55']='Im der Abbildung 1 wird die Abweichung über die Laststufen in % dargestellt'
ws['A55'].font=ft

img = Image('Abweichung_1.1.png')
img.height = 375
img.width= 600
img.anchor = 'A57'
ws.add_image(img)

ws['B77']='Abbildung 1'
ws['B77'].font=ft2

ws['A79']='Die Genauigkeit beträgt:'
ws['A79'].font=ft3

ws['E79']=max(np.round(abs(Genauigkeit),decimals=2))
ws['E79'].font=ft3

ws['F79']='[L/Min]'
ws['F79'].font=ft3


# Speichern des Excels

wb.save('Kalibrierung Wasser 1.1.xlsx')
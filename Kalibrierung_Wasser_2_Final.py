"""
Programm für die Kalibrierung von Wasser. Es handelt sich um die Version 2.
Es wird im Programm zusätzlich noch der Prüfbericht erstellt. Es werden dafür die Daten ins Excel exportiert.

Programm: Sandro Villiger
"""

# %% Import der Pakete

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# %% Veränderliche Variablen

# Pfad zu Datei mit den Rohdaten

file = r'C:\Users\Sandro Villiger\switchdrive\Arbeit IBI\Arbeit Uwe\037 Kallibrirung Killian\Kallibrirung der Wassermenge\2022-12-22-13_45_53.485_Kalibration_Wasser_2.csv'

U = 0.09

# %% Bestimmte Variablen

k = 2

# %% Import der Daten

data = pd.read_csv(file,sep=';')

Referenz=np.array(data['MW'])

Messung=np.array(data['Qw 20'])

Punkte= np.arange(0,len(Referenz),1)

# %% Plot der gesamten Daten
'''
plt.figure(figsize=(8, 5), dpi=300)
plt.plot(Punkte,Referenz,color='black',label='Referenz')
plt.plot(Punkte,Referenz,color='blue',label='Messung')
plt.grid()
plt.xticks(np.arange(0,10001,1000))
plt.legend()
plt.show()'''

# %% Slicing der Daten nach Zeitschriten

# Referenz Daten

Werte_1_r=Referenz[1750:1950]
Werte_2_r=Referenz[2200:2400]
Werte_3_r=Referenz[2600:2900]
Werte_4_r=Referenz[3150:3420]
Werte_5_r=Referenz[3680:4000]
Werte_6_r=Referenz[4220:4500]
Werte_7_r=Referenz[4800:5050]
Werte_8_r=Referenz[5350:5600]
Werte_9_r=Referenz[5940:6300]
Werte_10_r=Referenz[6580:6720]
Werte_11_r=Referenz[6950:7250]
Werte_12_r=Referenz[7500:7800]
Werte_13_r=Referenz[8000:8320]
Werte_14_r=Referenz[8550:8900]
Werte_15_r=Referenz[9100:9400]
Werte_16_r=Referenz[9650:10000]
Werte_17_r=Referenz[10220:10520]

# Prüfdaten Daten

Werte_1_m=Messung[1750:1950]
Werte_2_m=Messung[2200:2400]
Werte_3_m=Messung[2600:2900]
Werte_4_m=Messung[3150:3420]
Werte_5_m=Messung[3680:4000]
Werte_6_m=Messung[4220:4500]
Werte_7_m=Messung[4800:5050]
Werte_8_m=Messung[5350:5600]
Werte_9_m=Messung[5940:6300]
Werte_10_m=Messung[6580:6720]
Werte_11_m=Messung[6950:7250]
Werte_12_m=Messung[7500:7800]
Werte_13_m=Messung[8000:8320]
Werte_14_m=Messung[8550:8900]
Werte_15_m=Messung[9100:9400]
Werte_16_m=Messung[9650:10000]
Werte_17_m=Messung[10220:10520]

# %% Plot der Daten mit den verwendeten Messdaten
'''
plt.figure(figsize=(8, 5), dpi=300)
plt.plot(Punkte,Referenz,color='black',label='Referenz')
plt.plot(Punkte,Referenz,color='blue',label='Messung')


plt.plot(np.arange(1750,1950),Werte_1_r,color='red',label='Auswertung')
plt.plot(np.arange(2200,2400),Werte_2_r,color='red')
plt.plot(np.arange(2600,2900),Werte_3_r,color='red')
plt.plot(np.arange(3150,3420),Werte_4_r,color='red')
plt.plot(np.arange(3680,4000),Werte_5_r,color='red')
plt.plot(np.arange(4220,4500),Werte_6_r,color='red')
plt.plot(np.arange(4800,5050),Werte_7_r,color='red')
plt.plot(np.arange(5350,5600),Werte_8_r,color='red')
plt.plot(np.arange(5940,6300),Werte_9_r,color='red')
plt.plot(np.arange(6580,6720),Werte_10_r,color='red')
plt.plot(np.arange(6950,7250),Werte_11_r,color='red')
plt.plot(np.arange(7500,7800),Werte_12_r,color='red')
plt.plot(np.arange(8000,8320),Werte_13_r,color='red')
plt.plot(np.arange(8550,8900),Werte_14_r,color='red')
plt.plot(np.arange(9100,9400),Werte_15_r,color='red')
plt.plot(np.arange(9650,10000),Werte_16_r,color='red')
plt.plot(np.arange(10220,10520),Werte_17_r,color='red')

plt.grid()
plt.xticks(np.arange(0,10001,1000))
plt.legend()
plt.show()
'''

#%% Berechnung der Mittelwerte


x1_R=Werte_1_r.mean()
x2_R=Werte_2_r.mean()
x3_R=Werte_3_r.mean()
x4_R=Werte_4_r.mean()
x5_R=Werte_5_r.mean()
x6_R=Werte_6_r.mean()
x7_R=Werte_7_r.mean()
x8_R=Werte_8_r.mean()
x9_R=Werte_9_r.mean()
x10_R=Werte_10_r.mean()
x11_R=Werte_11_r.mean()
x12_R=Werte_12_r.mean()
x13_R=Werte_13_r.mean()
x14_R=Werte_14_r.mean()
x15_R=Werte_15_r.mean()
x16_R=Werte_16_r.mean()
x17_R=Werte_17_r.mean()



x1_M=Werte_1_m.mean()
x2_M=Werte_2_m.mean()
x3_M=Werte_3_m.mean()
x4_M=Werte_4_m.mean()
x5_M=Werte_5_m.mean()
x6_M=Werte_6_m.mean()
x7_M=Werte_7_m.mean()
x8_M=Werte_8_m.mean()
x9_M=Werte_9_m.mean()
x10_M=Werte_10_m.mean()
x11_M=Werte_11_m.mean()
x12_M=Werte_12_m.mean()
x13_M=Werte_13_m.mean()
x14_M=Werte_14_m.mean()
x15_M=Werte_15_m.mean()
x16_M=Werte_16_m.mean()
x17_M=Werte_17_m.mean()

M1=np.array([x1_M,x2_M,x3_M,x4_M,x5_M,x6_M,x7_M,x8_M,x9_M])
M2=np.array([x17_M,x16_M,x15_M,x14_M,x13_M,x12_M,x11_M,x10_M,np.nan])

X_Stufen=np.array([x1_M,x2_M,x3_M,x4_M,x5_M,x6_M,x7_M,x8_M,x9_M,x10_M,x11_M,x12_M,x13_M,x14_M,x15_M,x16_M,x17_M])

#%% Berechnung der totalen Mittelwerte


X1R=(x1_R+x17_R)/2
X2R=(x2_R+x16_R)/2
X3R=(x3_R+x15_R)/2
X4R=(x4_R+x14_R)/2
X5R=(x5_R+x13_R)/2
X6R=(x6_R+x12_R)/2
X7R=(x7_R+x11_R)/2
X8R=(x8_R+x10_R)/2
X9R=x9_R


XR=np.array([X1R,X2R,X3R,X4R,X5R,X6R,X7R,X8R,X9R])


X1M=(x1_M+x17_M)/2
X2M=(x2_M+x16_M)/2
X3M=(x3_M+x15_M)/2
X4M=(x4_M+x14_M)/2
X5M=(x5_M+x13_M)/2
X6M=(x6_M+x12_M)/2
X7M=(x7_M+x11_M)/2
X8M=(x8_M+x8_M)/2
X9M=x9_M


XM=np.array([X1M,X2M,X3M,X4M,X5M,X6M,X7M,X8M,X9M])

#%% Auswertung

Laststufen=np.array([10,20,30,40,50,60,70,80,90])
print(XM-XR)
A = ((XM-XR ) / Laststufen) * 100

U_Wasser= U/100*Laststufen

Abweichung=XM - XR

Hysterese= abs(M1-M2)

Genauigkeit=abs(Abweichung)+2*abs(U_Wasser)

Messunsicherheit=np.concatenate((XR.reshape(-1,1),XM.reshape(-1,1),Abweichung.reshape(-1,1),Hysterese.reshape(-1,1),U_Wasser.reshape(-1,1),Genauigkeit.reshape(-1,1)),axis=1)

#%% Plot Auswertungsdiagramm


text=['-10%','-7.5%','-5%','-2.5%','0%','2.5%','5%','7.5%','10%']

plt.figure(figsize=(8, 5), dpi=300)
plt.plot(abs(Laststufen),A,color='green')
plt.plot(abs(Laststufen),A-k*(U),color='blue',linestyle='--')
plt.plot(abs(Laststufen),A+k*(U),color='blue',linestyle='--')
plt.hlines(0,abs(Laststufen)[0],abs(Laststufen)[-1],colors='black')
plt.hlines(10,abs(Laststufen)[0],abs(Laststufen)[-1],colors='red')
plt.hlines(-10,abs(Laststufen)[0],abs(Laststufen)[-1],colors='red')
plt.yticks(np.arange(-10,11,2.5), text)
plt.grid()
plt.savefig('Abweichung_2.png')
#plt.show()

#%% Excel Export

#%% Excel Export

wb=Workbook()

# Defieneren der Schrift

ft = Font(name='Verdana',size=10) # Standard
ft2 = Font(name='Verdana',size=8) # Bildbeschriftung
ft3 = Font(name='Verdana',size=10,bold=True) # Hervorgehoben

# Erstellen des Worksheets
ws=wb.active

# Grösse für der Zelle (Richtige grösse für das Drucken)
ws.column_dimensions['G'].width=10
ws.column_dimensions['B'].width=12
ws.column_dimensions['C'].width=13
ws.column_dimensions['D'].width=13
ws.column_dimensions['E'].width=17
ws.column_dimensions['F'].width=17
ws.column_dimensions['G'].width=13

# Fusszeile
ws.oddFooter.left.text = "Seite &[Page] von &N"
ws.oddFooter.left.size = 11
ws.oddFooter.left.font = "Verdana"

# Einführen des Logos
img = Image('HSLU_Logo.png')
img.height = 30
img.width= 200
img.anchor = 'F1'
ws.add_image(img)
img = Image('HSLU_Logo.png')
img.height = 30
img.width= 200
img.anchor = 'F50'
ws.add_image(img)


# Titel des Dokuments
ws['A1']='Kalibrierschein Wasser 1.2'
ws['A1'].font=Font(bold=True, underline='single',name='Verdana',size=16)

# Text für Prüfbestimmungen

ws['A4']='Kalibriergegenstand:'
ws['A4'].font=ft
ws['A5']='Typ:'
ws['A5'].font=ft
ws['A6']='Log.Nr.(STS 0209):'
ws['A6'].font=ft
ws['A7']='Serien NR:'
ws['A7'].font=ft
ws['A8']='Anzeigebereich:'
ws['A8'].font=ft
ws['A9']='Max. zul. Abweichung:'
ws['A9'].font=ft
ws['A11']='Referenznormal:'
ws['A11'].font=ft
ws['A12']='Typ:'
ws['A12'].font=ft
ws['A13']='Log.Nr.:'
ws['A13'].font=ft
ws['A14']='Serien NR:'
ws['A14'].font=ft
ws['A16']='Datum der Kalibrierung:'
ws['A16'].font=ft
ws['A17']='Prüfer:'
ws['A17'].font=ft

ws['E5']='Temperatur:'
ws['E6']='Rel. Luftfeuchte:'
ws['E7']='Umgebungsdruck:'
ws['E8']='Ort der Kalibrierung:'
ws['E5'].font=ft
ws['E6'].font=ft
ws['E7'].font=ft
ws['E8'].font=ft


# Abgrenzung
Liste=('A','B','C','D','E','F','G',)
for i in Liste:
    ws[f'{i}17'].border=Border( bottom=Side(style='thin'))

ws['A19']='Die Kalibrierung erfolgt nach der Verfahrensanweisung BAA08083 "Wasser" nach dem '
ws['A20']='Verfahren der DAkkS EA-4/02 M: 2013 "Ermittlung der Messunsicherheiten bei Kalibrierungen" '
ws['A19'].font=ft
ws['A20'].font=ft

# Text Mittelwerte
ws['A22']='Folgend werden die Mittelwerte über die verschieden Stufen für den Prüfkörper gezeigt'
ws['A22'].font=ft

ws['A24']='Mittelwerte des Prüfkörpers in Liter pro Minute'
ws['A24'].font=ft3


Tabelle_1=np.concatenate((XR.reshape(-1,1),M1.reshape(-1,1),M2.reshape(-1,1)),axis=1)
Tabelle_1_Export=pd.DataFrame(np.round(Tabelle_1,decimals=2))
for r in dataframe_to_rows(Tabelle_1_Export):
    ws.append(r)

for row in ws['A25:A35']:
    for cell in row:
        cell.value = None

for row in ws['A25:H25']:
    for cell in row:
        cell.value = None

ws.move_range("B27:D35", rows=0, cols=-1)

for i in range(27,36,1):
    ws[f'A{i}'].font=Font(bold=True)
    ws[f'A{i}'].fill=PatternFill("solid", fgColor="e6e6e6")

ws['A26'].alignment = Alignment(wrap_text=True,vertical="top")
ws['A26'].value='Referenz'
ws['A26'].font=Font(bold=True)
ws['A26'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['B26'].alignment = Alignment(wrap_text=True,vertical="top")
ws['B26'].value='M1'
ws['B26'].font=Font(bold=True)
ws['B26'].fill=PatternFill("solid", fgColor="e6e6e6")

ws['C26'].alignment = Alignment(wrap_text=True,vertical="top")
ws['C26'].value='M2'
ws['C26'].font=Font(bold=True)
ws['C26'].fill=PatternFill("solid", fgColor="e6e6e6")

Liste1=('A','B','C')

for L in Liste1:
    for i in range(26,36,1):
        ws[f'{L}{i}'].border=Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

ws['A37']='Messunsicherheit'
ws['A37'].font=ft3

Messunsicherheit_export=pd.DataFrame(Messunsicherheit)

for r in dataframe_to_rows(np.round(Messunsicherheit_export,decimals=2)):
    ws.append(r)

for row in ws['A38:A48']:
    for cell in row:
        cell.value = None

for row in ws['B38:G38']:
    for cell in row:
        cell.value = None

ws.move_range("B40:G48", rows=0, cols=-1)

ws['A39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['A39'].value='Referenz\n[L/Min]'
ws['A39'].font=Font(bold=True)
ws['A39'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['B39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['B39'].value='Mittelwert\n[L/Min]'
ws['B39'].font=Font(bold=True)
ws['B39'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['C39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['C39'].value='Abweichung\n[L/Min]'
ws['C39'].font=Font(bold=True)
ws['C39'].fill=PatternFill("solid", fgColor="e6e6e6")


ws['D39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['D39'].value='Hysterese\n[L/Min]'
ws['D39'].font=Font(bold=True)
ws['D39'].fill=PatternFill("solid", fgColor="e6e6e6")

ws['E39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['E39'].value='Messunsicherheit\n[L/Min]'
ws['E39'].font=Font(bold=True)
ws['E39'].fill=PatternFill("solid", fgColor="e6e6e6")

ws['F39'].alignment = Alignment(wrap_text=True,vertical="top")
ws['F39'].value='Genauigkeit\nSensor\n[L/Min]'
ws['F39'].font=Font(bold=True,size=11)
ws['F39'].fill=PatternFill("solid", fgColor="e6e6e6")

Liste1=('A','B','C','D','E','F')

for L in Liste1:
    for i in range(39,49,1):
        ws[f'{L}{i}'].border=Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

ws['A53']='Die Abweichung des Referenznormals beträgt:'
ws['A53'].font=ft

ws['E53']=U
ws['E53'].font=ft3

ws['F53']='%'
ws['F53'].font=ft3

ws['A55']='Im der Abbildung 1 wird die Abweichung über die Laststufen in % dargestellt'
ws['A55'].font=ft

img = Image('Abweichung_2.png')
img.height = 375
img.width= 600
img.anchor = 'A57'
ws.add_image(img)

ws['B77']='Abbildung 1'
ws['B77'].font=ft2

ws['A79']='Die Genauigkeit beträgt:'
ws['A79'].font=ft3

ws['E79']=max(np.round(abs(Genauigkeit),decimals=2))
ws['E79'].font=ft3

ws['F79']='[L/Min]'
ws['F79'].font=ft3

wb.save('Kalibrierung Wasser 2.xlsx')
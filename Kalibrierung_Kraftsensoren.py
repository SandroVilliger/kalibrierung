"""
Programm für die Kalibrierung Kraftsensoren

Programm: Sandro Villiger
"""

#%% Pakete

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.dataframe import dataframe_to_rows

#%%



# Pfad zu datei und Import der Datei

file = r'Testlauf_2_Kraftgeregelt.XLSX'

# Header Name

name = ['Zeit', 'Referenz kN', 'Prüfling kN']

# Import der Daten

data = pd.read_excel(file, header=None, names=name)

# Löschen der unnötigen Daten

delet = np.arange(0, 49, 1)


Werte = data.drop(delet, axis=0)

# Einlesen der Daten in np.array

Zeit = np.round(np.array(Werte['Zeit'], dtype='float64'), 1)
Referenz = np.array(Werte['Referenz kN'], dtype='float64')
Prüfling = np.array(Werte['Prüfling kN'], dtype='float64')

'''
Zugverscuch = 1
Druckversuch = 0
'''

Versuch = 0

# maximale Prüfkraft

F_max = 2

if Versuch == 0:
    F_max = F_max * -1
else:
    F_max = F_max * 1

# Messstrom

S=2

# Prüfstufen


y_Beschriftung = np.linspace(0, F_max, num=5)
Stufen = np.array([0.125, 0.25, 0.375, 0.5, 0.675, 0.75, 0.875, 1])

plt.figure(figsize=(8, 5), dpi=300)
plt.plot(Zeit,Referenz,color='black')

plt.hlines(0,min(Zeit),max(Zeit),colors='grey')
plt.vlines(min(Zeit),0,F_max,colors='grey')
plt.title('Versuchsablauf Referenznormal', pad=12,fontsize=12)
plt.xlabel('Zeit [s]' ,fontsize=12,)
plt.ylabel('Kraft [kN]' ,fontsize=12,)
plt.yticks(y_Beschriftung)
plt.savefig('Gesamte Messung.png')
#plt.show()

Laststufen = F_max * Stufen
Mittelwerte_Refernz = [[], [], [], [], [], [], [],  []]

Mittelwerte_Prüfling = [[], [], [], [], [], [], [], []]

Null_Last=Werte_F = Werte[(Werte['Referenz kN'] >= 0- 0.01) & (Werte['Referenz kN'] <=0 + 0.01)]['Referenz kN'].mean()

for x in range(len(Laststufen)):
    Werte_F = Werte[(Werte['Referenz kN'] >= Laststufen[x] - 0.01) & (Werte['Referenz kN'] <= Laststufen[x] + 0.01)]
    Werte_F_index = Werte[(Werte['Referenz kN'] >= Laststufen[x] - 0.01) & (Werte['Referenz kN'] <= Laststufen[x] + 0.01)].index.values
    res = np.where(Werte_F_index[:-1] + 1 != Werte_F_index[1:])[0]
    #print(res)
    res = res[res > 10]

    X_max_1 = Werte_F[10:res[0] -9]
    X_max_2 = Werte_F[res[0] + 10:res[1] -9]
    X_max_3 = Werte_F[res[1] + 10:res[2] -9]
    X_max_4 = Werte_F[res[2] + 10:-10]
    Mittelwert = X_max_4.mean()
    #print(x)
    #print(res)

    Mittelwerte_Refernz[x].append(X_max_1.mean()[1])
    Mittelwerte_Refernz[x].append(X_max_2.mean()[1])
    Mittelwerte_Refernz[x].append(X_max_3.mean()[1])
    Mittelwerte_Refernz[x].append(X_max_4.mean()[1])

    Mittelwerte_Prüfling[x].append(X_max_1.mean()[2])
    Mittelwerte_Prüfling[x].append(X_max_2.mean()[2])
    Mittelwerte_Prüfling[x].append(X_max_3.mean()[2])
    Mittelwerte_Prüfling[x].append(X_max_4.mean()[2])
    '''plt.figure(figsize=(8, 5), dpi=300)
    plt.plot(Zeit, Referenz, color='black')
    plt.plot(X_max_1['Zeit'], X_max_1['Referenz kN'], color='blue')
    plt.plot(X_max_2['Zeit'], X_max_2['Referenz kN'], color='red')
    plt.plot(X_max_3['Zeit'], X_max_3['Referenz kN'], color='blue')
    plt.plot(X_max_4['Zeit'], X_max_4['Referenz kN'], color='red')
    plt.vlines(min(Zeit), 0, F_max, colors='grey')
    plt.title('Versuchsablauf Referenznormal', pad=12, fontsize=12)
    plt.xlabel('Zeit [s]', fontsize=12, )
    plt.ylabel('Kraft [kN]', fontsize=12, )
    plt.yticks(y_Beschriftung)
    plt.show()'''

Mittelwerte_Refernz = np.array(Mittelwerte_Refernz)
Mittelwerte_Prüfling = np.array(Mittelwerte_Prüfling)


Refernz_mean = np.mean(Mittelwerte_Refernz, axis=1)
Prüfling_mean = np.mean(Mittelwerte_Prüfling, axis=1)

A = np.array(((Prüfling_mean-Refernz_mean ) / Laststufen) * 100)

# abweichung Referenznormal
Urn= 0.022


# Abweichung Messverstärker

Umv = 0.05

# Abweichung

U=np.sqrt((Urn)**2+(Umv)**2)

#print(U)

k=2

text = ['F1', 'F2','F3', 'F4', 'F5', 'F6', 'F7','F8']

plt.figure(figsize=(8, 5), dpi=300)
plt.plot(abs(Laststufen),A,color='green')
plt.plot(abs(Laststufen),A-k*(U),color='blue',linestyle='--')
plt.plot(abs(Laststufen),A+k*(U),color='blue',linestyle='--')
plt.hlines(0.2,abs(Laststufen)[0],abs(Laststufen)[-1],colors='red')
plt.hlines(-0.2,abs(Laststufen)[0],abs(Laststufen)[-1],colors='red')
plt.xticks(abs(Laststufen), text)
plt.grid()
plt.savefig('Abweichung.png')
#plt.show()


'''
Ausgabe für Fehler korrekter 
'''

Korrektur_max= (S/ ((F_max*1000)-(A[-1]/100)*F_max*1000))*F_max*1000

Korrektur_unten=((A[0]/100)*F_max*1000)+(F_max*1000*Stufen[0])

Korrektur_min=(Stufen[0]*S*F_max*1000/Korrektur_unten)-2

print(0,'N','=', Korrektur_min,'mV/V')
print(-F_max*1000,'N','=', Korrektur_max,'mV/V')


#%% Export in Excel

# Erstellen des Excel

wb=Workbook()

# Defineren der Schrift
ft = Font(name='Verdana',size=10)
ft2 = Font(name='Verdana',size=8)
ft3 = Font(name='Verdana',size=10,bold=True)

# Erstellen des Worksheets
ws=wb.active

# Name des Worksheets
ws.title = 'Kalibrierung'

# Grösse für der Zelle (Richtige gröse für das Drucken)
ws.column_dimensions['G'].width=16
ws.column_dimensions['Q'].width=16

# Fusszeile
ws.oddFooter.left.text = "Seite &[Page] von &N"
ws.oddFooter.left.size = 11
ws.oddFooter.left.font = "Verdana"

# Einführen des Logos
img = Image('HSLU_Logo.png')
img.height = 30
img.width= 200
img.anchor = 'H1'
ws.add_image(img)
img = Image('HSLU_Logo.png')
img.height = 30
img.width= 200
img.anchor = 'H52'
ws.add_image(img)
img = Image('HSLU_Logo.png')
img.height = 30
img.width= 200
img.anchor = 'H103'
ws.add_image(img)

# Titel des Dokuments
ws['A1']='Kalibrierschein'
ws['A1'].font=Font(bold=True, underline='single',name='Verdana',size=16)

# Text für Prüfbestimmungen

ws['A4']='Kalibriergegenstand:'
ws['A4'].font=ft
ws['A5']='Typ:'
ws['A5'].font=ft
ws['A6']='Log.Nr.(STS 0209):'
ws['A6'].font=ft
ws['A7']='Serien NR:'
ws['A7'].font=ft
ws['A8']='Anzeigebereich:'
ws['A8'].font=ft
ws['A9']='Max. zul. Abweichung'
ws['A9'].font=ft
ws['A11']='Referenznormal:'
ws['A11'].font=ft
ws['A12']='Typ:'
ws['A12'].font=ft
ws['A13']='Log.Nr.:'
ws['A13'].font=ft
ws['A14']='Serien NR:'
ws['A14'].font=ft
ws['A16']='Datum der Kalibrierung:'
ws['A16'].font=ft
ws['A17']='Prüfer:'
ws['A17'].font=ft

ws['F5']='Temperatur:'
ws['F6']='Rel. Luftfeuchte:'
ws['F7']='Umgebungsdruck:'
ws['F8']='Ort der Kalibrierung:'
ws['F5'].font=ft
ws['F6'].font=ft
ws['F7'].font=ft
ws['F8'].font=ft

# Abgrenzung
Liste=('A','B','C','D','E','F','G','H','I','J')
for i in Liste:
    ws[f'{i}18'].border=Border( bottom=Side(style='thin'))



ws['A20']='Die Kalibrierung erfolgt nach der Verfahrensanweisung BAA08082 "Kalibration Kraft" nach dem '
ws['A21']='Verfahren der DAkkS EA-4/02 M: 2013"Ermittlung der Messunsicherheiten bei Kalibrierungen" '
ws['A20'].font=ft
ws['A21'].font=ft

ws['A23']='In der Abbildung 1 ist die gesamte Messung dargestellt.'
ws['A23'].font=ft

# Import des Bildes der gesamten Zeitreihe

img = Image('Gesamte Messung.png')
img.height = 375
img.width= 600
img.anchor = 'B25'
ws.add_image(img)

# Bild Beschriftung

ws['C45']='Abbildung 1'
ws['C45'].font=ft2

# Text Mittelwerte
ws['A55']='Folgend werden die Mittelwerte über die verschieden Laststufen gezeigt'
ws['A55'].font=ft

ws['A57']='Mittelwerte der Referenzmessung'
ws['A57'].font=ft3



# Export der Mittelwerte

Mittelwerte_Refernz_Export=pd.DataFrame(Mittelwerte_Refernz,columns=['M1','M2','M3','M4'],index=['F1','F2','F3','F4','F5','F6','F7','FMax'])

for r in dataframe_to_rows(np.round(Mittelwerte_Refernz_Export,decimals=5),index=True,header=True):
    ws.append(r)
ws.move_range("A58:F58", rows=1, cols=0)
ws.move_range("A59:F68", rows=1, cols=1)
ws['C60'].font=Font(bold=True)
ws['C60'].fill=PatternFill("solid", fgColor="e6e6e6")
ws['D60'].font=Font(bold=True)
ws['D60'].fill=PatternFill("solid", fgColor="e6e6e6")
ws['E60'].font=Font(bold=True)
ws['E60'].fill=PatternFill("solid", fgColor="e6e6e6")
ws['F60'].font=Font(bold=True)
ws['F60'].fill=PatternFill("solid", fgColor="e6e6e6")
for i in range(60,69,1):
    ws[f'B{i}'].font=Font(bold=True)
    ws[f'B{i}'].fill=PatternFill("solid", fgColor="e6e6e6")

Liste1=('B','C','D','E','F')

for L in Liste1:
    for i in range(60,69,1):
        ws[f'{L}{i}'].border=Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


ws['A71']='Mittelwerte des Prüfkörpers'
ws['A71'].font=ft3

Mittelwerte_Prüfling_Export=pd.DataFrame(np.round(Mittelwerte_Prüfling,decimals=5),columns=['M1','M2','M3','M4'],index=['F1','F2','F3','F4','F5','F6','F7','FMax'])

for r in dataframe_to_rows(Mittelwerte_Prüfling_Export,index=True,header=True):
    ws.append(r)
ws.move_range("A72:F72", rows=1, cols=0)
ws.move_range("A73:F81", rows=1, cols=1)
ws['C74'].font=Font(bold=True)
ws['C74'].fill=PatternFill("solid", fgColor="e6e6e6")
ws['D74'].font=Font(bold=True)
ws['D74'].fill=PatternFill("solid", fgColor="e6e6e6")
ws['E74'].font=Font(bold=True)
ws['E74'].fill=PatternFill("solid", fgColor="e6e6e6")
ws['F74'].font=Font(bold=True)
ws['F74'].fill=PatternFill("solid", fgColor="e6e6e6")
for i in range(74,83,1):
    ws[f'B{i}'].font=Font(bold=True)
    ws[f'B{i}'].fill=PatternFill("solid", fgColor="e6e6e6")



for L in Liste1:
    for i in range(74,83,1):
        ws[f'{L}{i}'].border=Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Auswertung

ws['A85']='In der folgenden Tabelle werden die Resultate der Mittelwert und die prozentualen '
ws['A86']='Abweichungen gezeigt'
ws['A85'].font=ft
ws['A86'].font=ft

i=Refernz_mean.reshape(8,1)
i=np.append(i,Prüfling_mean.reshape(8,1),axis=1)
i=np.append(i,A.reshape(8,1),axis=1)


Auswertung=pd.DataFrame(i,columns=['Referenz','Prüfling','A'],index=['F1','F2','F3','F4','F5','F6','F7','FMax'])


for r in dataframe_to_rows(np.round(Auswertung,decimals=5),index=True,header=True):
    ws.append(r)

ws.move_range("A87:F87", rows=1, cols=0)
ws.move_range("A88:F96", rows=1, cols=1)

ws['C89'].font=Font(bold=True,size=10)
ws['C89'].fill=PatternFill("solid", fgColor="e6e6e6")
ws['D89'].font=Font(bold=True,size=10)
ws['D89'].fill=PatternFill("solid", fgColor="e6e6e6")
ws['E89'].font=Font(bold=True,size=10)
ws['E89'].fill=PatternFill("solid", fgColor="e6e6e6")

for i in range(89,98,1):
    ws[f'B{i}'].font=Font(bold=True)
    ws[f'B{i}'].fill=PatternFill("solid", fgColor="e6e6e6")

for i in range(90,98,1):
    ws[f'E{i}'].fill=PatternFill("solid", fgColor="adebad")


Liste2=('B','C','D','E',)

for L in Liste2:
    for i in range(89,98,1):
        ws[f'{L}{i}'].border=Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


# abweichung Referenznormal
Urn= 0.022


# Abweichung Messverstärker

Umv = 0.05

ws['A104']='Die Abweichung des Referenznormals beträgt:'
ws['A104'].font=ft

ws['F104']=Urn
ws['F104'].font=ft3

ws['A106']='Die Abweichung des Messverstärkers beträgt:'
ws['A106'].font=ft

ws['F106']=Umv
ws['F106'].font=ft3

ws['A109']='Im der Abbildung 2 wird die Abweichung dargestellt'
ws['A109'].font=ft

img = Image('Abweichung.png')
img.height = 375
img.width= 600
img.anchor = 'B110'
ws.add_image(img)

# Bild Beschriftung

ws['C130']='Abbildung 2'
ws['C130'].font=ft2




# Speichern des Excels

wb.save('Kalibrierung.xlsx')